import streamlit as st
import pandas as pd
from datetime import datetime
from streamlit_drawable_canvas import st_canvas
import base64
from io import BytesIO
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import io
import firebase_admin
from firebase_admin import credentials, firestore

# Page configuration
st.set_page_config(
    page_title="안전작업허가서",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Data storage
EXCEL_TEMPLATE = "안전작업허가서.xlsx"

# Initialize Firebase
try:
    # Try to load from JSON file first (for local development)
    json_key_path = "home-assistant-7430-firebase-adminsdk-9qngt-c2f6659ad5.json"
    if os.path.exists(json_key_path):
        cred = credentials.Certificate(json_key_path)
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
        db = firestore.client(database_id='default')
    else:
        # Fallback to secrets.toml (for deployment)
        firebase_creds = {
            "type": st.secrets.firebase.type,
            "project_id": st.secrets.firebase.project_id,
            "private_key_id": st.secrets.firebase.private_key_id,
            "private_key": st.secrets.firebase.private_key.replace('\\n', '\n'),
            "client_email": st.secrets.firebase.client_email,
            "client_id": st.secrets.firebase.client_id,
            "auth_uri": st.secrets.firebase.auth_uri,
            "token_uri": st.secrets.firebase.token_uri,
            "auth_provider_x509_cert_url": st.secrets.firebase.auth_provider_x509_cert_url,
            "client_x509_cert_url": st.secrets.firebase.client_x509_cert_url,
        }
        if not firebase_admin._apps:
            cred = credentials.Certificate(firebase_creds)
            firebase_admin.initialize_app(cred)
        db = firestore.client(database_id='default')
except Exception as e:
    st.error(f"Firebase initialization error: {e}")
    db = None

SAFETY_CHECK_CATEGORY_NAMES = {
    '첨부서류',
    '안전조치 요구사항',
    '밀폐공간',
    '정전',
    '굴착',
    '고소',
    '중장비',
}


def normalize_clock_time(value):
    """사용자가 입력한 시각을 HH:MM 형식으로 정규화한다."""
    value = str(value or "").strip()
    if not value:
        return "", None

    if value.isdigit():
        if len(value) <= 2:
            hour, minute = int(value), 0
        elif len(value) == 3:
            hour, minute = int(value[0]), int(value[1:])
        elif len(value) == 4:
            hour, minute = int(value[:2]), int(value[2:])
        else:
            return value, "숫자는 최대 4자리로 입력해 주세요."
    else:
        parts = value.split(":")
        if len(parts) != 2 or not all(part.isdigit() for part in parts):
            return value, "숫자 또는 HH:MM 형식으로 입력해 주세요."
        hour, minute = map(int, parts)

    if not 0 <= hour <= 23 or not 0 <= minute <= 59:
        return value, "올바른 시간 범위(00:00~23:59)로 입력해 주세요."

    return f"{hour:02d}:{minute:02d}", None


def normalize_time_widget(key):
    """시간 입력이 끝나면 해당 위젯 값을 HH:MM 형식으로 바꾼다."""
    normalized, error = normalize_clock_time(st.session_state.get(key, ""))
    if error:
        st.session_state[f"{key}_error"] = error
        return

    st.session_state[key] = normalized
    st.session_state.pop(f"{key}_error", None)


def format_time_value(value):
    """datetime/time/string 값을 Excel에 쓰기 좋은 HH:MM 문자열로 맞춘다."""
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")

    value = str(value).strip()
    if len(value) >= 5 and value[2] == ":":
        return value[:5]
    return value


def format_worker_count(value):
    """작업자 명수를 Excel 작업자 칸에 표시할 문자열로 변환한다."""
    if value in (None, ""):
        return ""

    try:
        count = int(value)
    except (TypeError, ValueError):
        return str(value).strip()

    if count <= 0:
        return ""
    return f"{count}명"


def decode_form_signatures(form_data):
    """Excel 생성에 필요한 제출 건의 서명만 지연 디코딩한다."""
    encoded_signatures = form_data.get('signatures')
    if not isinstance(encoded_signatures, dict):
        return form_data

    decoded_signatures = {}
    for sig_key, sig_value in encoded_signatures.items():
        if not isinstance(sig_value, str):
            decoded_signatures[sig_key] = sig_value
            continue

        try:
            import numpy as np

            if sig_value.startswith('data:image/png;base64,'):
                sig_value = sig_value.split(',', 1)[1]
            img_bytes = base64.b64decode(sig_value)
            img = PILImage.open(io.BytesIO(img_bytes))
            decoded_signatures[sig_key] = np.array(img)
        except Exception as e:
            st.error(f"Error decoding signature {sig_key}: {e}")

    decoded_form = dict(form_data)
    decoded_form['signatures'] = decoded_signatures
    return decoded_form


def load_submitted_forms():
    if db is None:
        return []
    try:
        forms_ref = db.collection('forms')
        docs = forms_ref.order_by('timestamp', direction=firestore.Query.DESCENDING).stream()
        forms = []
        for doc in docs:
            form_data = doc.to_dict()
            # 관리자 수정 시 조건 검색을 거치지 않고 해당 문서를 바로 갱신한다.
            form_data['_doc_id'] = doc.id
            
            # Reconstruct safety_checks from flattened format
            safety_checks = {}
            for k, v in form_data.items():
                # 저장 시 평탄화한 실제 안전점검 항목만 다시 묶는다.
                # gas_measurements 같은 일반 필드의 밑줄은 안전점검 항목으로 오인하지 않는다.
                for category in SAFETY_CHECK_CATEGORY_NAMES:
                    prefix = f"{category}_"
                    if k.startswith(prefix):
                        check_name = k[len(prefix):]
                        safety_checks.setdefault(category, {})[check_name] = v
                        break
            
            if safety_checks:
                form_data['safety_checks'] = safety_checks
            
            forms.append(form_data)
        return forms
    except Exception as e:
        st.error(f"Error loading forms: {e}")
        return []

def save_form(form_data):
    if db is None:
        return False
    try:
        # Flatten nested data to avoid Firestore nested entity errors
        form_data_to_save = {}
        
        # Copy simple fields (exclude signatures for now)
        for k, v in form_data.items():
            if k not in ['signatures', 'safety_checks']:
                form_data_to_save[k] = v
        
        # Flatten safety_checks
        if 'safety_checks' in form_data:
            for category, checks in form_data['safety_checks'].items():
                if isinstance(checks, dict):
                    for check_name, check_value in checks.items():
                        form_data_to_save[f"{category}_{check_name}"] = check_value
        
        # Convert signatures to base64 strings
        if 'signatures' in form_data:
            signature_base64 = {}
            for sig_key, sig_data in form_data['signatures'].items():
                if sig_data is not None and hasattr(sig_data, 'tolist'):
                    try:
                        import base64
                        import numpy as np
                        # Convert numpy array to image
                        img_array = np.array(sig_data)
                        img = PILImage.fromarray(img_array.astype('uint8'))
                        
                        # Convert to bytes
                        img_bytes = io.BytesIO()
                        img.save(img_bytes, format='PNG')
                        img_bytes.seek(0)
                        
                        # Convert to base64
                        img_base64 = base64.b64encode(img_bytes.getvalue()).decode('utf-8')
                        signature_base64[sig_key] = f"data:image/png;base64,{img_base64}"
                    except Exception as e:
                        st.error(f"Error encoding signature {sig_key}: {e}")
                else:
                    st.warning(f"Signature {sig_key} data is None or not a numpy array")
            
            if signature_base64:
                form_data_to_save['signatures'] = signature_base64
        
        # Add metadata
        form_data_to_save['timestamp'] = datetime.now()
        
        db.collection('forms').add(form_data_to_save)
        return True
    except Exception as e:
        st.error(f"Error saving form: {e}")
        return False

def delete_form(form_id):
    if db is None:
        return False
    try:
        forms_ref = db.collection('forms')
        docs = forms_ref.where('id', '==', form_id).stream()
        for doc in docs:
            doc.reference.delete()
        return True
    except Exception as e:
        st.error(f"Error deleting form: {e}")
        return False


def delete_all_forms():
    """관리자 페이지의 전체 삭제용: Firestore forms 컬렉션 문서를 모두 삭제한다."""
    if db is None:
        return None
    try:
        docs = list(db.collection('forms').stream())
        deleted_count = 0
        batch = db.batch()

        for index, doc in enumerate(docs, start=1):
            batch.delete(doc.reference)
            deleted_count += 1
            if index % 450 == 0:
                batch.commit()
                batch = db.batch()

        if deleted_count % 450:
            batch.commit()

        return deleted_count
    except Exception as e:
        st.error(f"Error deleting all forms: {e}")
        return None


def update_manager_name(form_id, manager_name, doc_id=None):
    """관리자 페이지에서 제출된 건에 담당자(관리자)를 지정/변경한다."""
    return update_form_fields(form_id, {'manager_name': manager_name}, doc_id=doc_id)

def update_form_fields(form_id, updates, doc_id=None):
    """관리자 페이지에서 제출된 건의 특정 필드(들)를 지정/변경한다."""
    if db is None:
        return False
    try:
        if doc_id:
            db.collection('forms').document(doc_id).update(updates)
            return True

        # 이전 데이터처럼 문서 ID가 없는 경우에만 기존 검색 방식을 사용한다.
        forms_ref = db.collection('forms')
        docs = forms_ref.where('id', '==', form_id).stream()
        updated = False
        for doc in docs:
            doc.reference.update(updates)
            updated = True
        return updated
    except Exception as e:
        st.error(f"Error updating form: {e}")
        return False

# 실제 "안전작업허가서.xlsx" 병합 셀 구조를 스캔해서 확인한 정확한 좌표.
# key: (카테고리, 체크리스트 항목명) -> 체크박스 셀 주소
# "보충작업허가"는 서식에 대응하는 체크박스 자체가 없어 매핑하지 않음(엑셀 미반영).
# 밀폐공간/정전은 카테고리 전체 체크 + 위치선택 로직이 있어 아래 fill_excel_template에서 별도 처리함.
CHECKBOX_CELL_MAP = {
    ('첨부서류', '작업계획서'): 'K10',
    ('첨부서류', '소화기목록'): 'T10',
    ('첨부서류', '특수작업절차서'): 'AD10',
    ('첨부서류', '기술자료(도면)'): 'K11',
    ('첨부서류', '안전장구 목록'): 'T11',
    ('첨부서류', '굴착도면'): 'AD11',

    ('안전조치 요구사항', '작업구역 설정(출입경고 표지)'): 'S13',
    ('안전조치 요구사항', '작업구역 가연성물질 제거'): 'S14',
    ('안전조치 요구사항', '밸브차단 및 차단표지부착(도면 비교)'): 'S15',
    ('안전조치 요구사항', '맹판설치 및 표지부착(도면 비교)'): 'S16',
    ('안전조치 요구사항', '위험물질(가연성분진 포함)방출 및 처리'): 'S17',
    ('안전조치 요구사항', '용기개방 및 압력방출'): 'AJ13',
    ('안전조치 요구사항', '용기내부 세정 및 처리'): 'AJ14',
    ('안전조치 요구사항', '불활성가스 치환 및 환기'): 'AJ15',
    ('안전조치 요구사항', '환기장비'): 'AJ16',
    ('안전조치 요구사항', '가스농도 측정'): 'AJ17',
    ('안전조치 요구사항', '조명장비'): 'AW13',
    ('안전조치 요구사항', '소화기'): 'AW14',
    ('안전조치 요구사항', '안전장구'): 'AW15',
    ('안전조치 요구사항', '안전교육'): 'AW16',
    ('안전조치 요구사항', '운전요원의 입회'): 'AW17',

    ('밀폐공간', '통신수단'): 'L19',
    ('밀폐공간', '구명장구(줄, 송기마스크)'): 'AB19',

    ('굴착', '가스,기계,소방배관'): 'S29',
    ('굴착', '전기,계장,통신'): 'S30',

    ('고소', '작업발판, 안전난간'): 'S31',
    ('고소', '안전사다리 사용'): 'S32',
    ('고소', '안전대 착용·부착'): 'AJ31',
    ('고소', '추락방지망'): 'BA31',

    ('중장비', '기상, 노면상태'): 'O34',
    ('중장비', '자격증소지'): 'AJ33',
    ('중장비', '현장책임자 감독'): 'BA33',
    ('중장비', '전선, 설비 간섭'): 'AA34',
    ('중장비', '신호수배치'): 'AK34',
    ('중장비', '매트 등 부속장구'): 'BA34',
}

# 체크박스가 아니라 텍스트(이름/시간 등)를 채워 넣는 항목의 셀 주소
TEXT_FIELD_CELL_MAP = {
    '전원복구 요청자': 'P28',
    '전원복구 시간': 'AB28',
    '투입장비명': 'L33',
    '운전원 성명': 'L35',
    '굴착 가스,기계,소방배관 점검자': 'AB29',
    '굴착 전기,계장,통신 점검자': 'AB30',
    # 실제 서식엔 "허가기간" 칸이 굴착 첫 행(29행)에만 있어 굴착 전체 공용으로 사용
    '굴착 허가기간': 'AT29',
    '고소 허가기간': 'AB32',
    '중장비 허가기간': 'AB35',
    '밀폐공간 허가기간': 'AK19',
    # 실제 서식엔 "허가기간" 칸이 정전 첫 행(제어실, 25행)에만 있어 정전 전체 공용으로 사용
    '정전 허가기간': 'AT25',
}

# 밀폐공간 가스농도 측정값 4건: Excel 서식의 22~23행 좌/우 입력 묶음에 대응한다.
GAS_MEASUREMENT_CELL_MAP = [
    {
        'material_name': 'E22',
        'result': 'M22',
        'measurement_time': 'Q22',
        'measurer_confirmer': 'W22',
    },
    {
        'material_name': 'AE22',
        'result': 'AM22',
        'measurement_time': 'AQ22',
        'measurer_confirmer': 'AW22',
    },
    {
        'material_name': 'E23',
        'result': 'M23',
        'measurement_time': 'Q23',
        'measurer_confirmer': 'W23',
    },
    {
        'material_name': 'AE23',
        'result': 'AM23',
        'measurement_time': 'AQ23',
        'measurer_confirmer': 'AW23',
    },
]

# 밀폐공간/정전 카테고리 전체("해당사항") 체크박스
CATEGORY_CHECKBOX_CELL_MAP = {
    '밀폐공간': 'E19',
    '정전': 'E24',
}

# 정전: 제어실/현장 중 선택한 위치에 따라 체크박스 셀이 달라짐
POWER_OUTAGE_LOCATION_CELLS = {
    '제어실': {'스위치/차단기 내림': 'U25', '잠금장치 시건, 표지부착': 'AK25'},
    '현장': {'스위치/차단기 내림': 'U26', '잠금장치 시건, 표지부착': 'AK26'},
}

# 위험성평가 필요 여부: 작업절차서변화/작업상이 각각 유·무에 따른 체크박스 셀
RISK_ASSESSMENT_CELLS = {
    'risk_assessment_change': {'유': 'AV10', '무': 'AZ10'},
    'risk_assessment_diff': {'유': 'AV11', '무': 'AZ11'},
}

# 작업종류 선택 시 해당 작업 구역의 대표 체크박스에 표시한다.
# "용접작업"과 "기타"는 현재 Excel 템플릿에 별도 대표 체크칸이 없다.
WORK_TYPE_CHECKBOX_CELL_MAP = {
    '밀폐작업': 'E19',
    '전기작업': 'E24',
    '굴착작업': 'E29',
    '고소작업': 'E31',
    '중장비작업': 'E33',
}


def fill_excel_template(form_data):
    if not os.path.exists(EXCEL_TEMPLATE):
        return None

    # 관리자 목록을 표시할 때는 무거운 서명 변환을 하지 않고,
    # 실제로 이 제출 건을 다운로드할 때만 변환한다.
    form_data = decode_form_signatures(form_data)
    
    wb = load_workbook(EXCEL_TEMPLATE)
    ws = wb.active
    
    safety_data = form_data.get('safety_checks', {})
    text_field_data = form_data.get('safety_text_fields', {})
    completion_time_value = format_time_value(
        form_data.get('permit_end_time') or form_data.get('completion_time')
    )
    worker_count_value = format_worker_count(form_data.get('worker_count'))
    
    cell_mapping = {
        # 허가번호 = 공사명 + 순번(예: 공사명01) - top-left of I3:AD3 merged range
        'I3': form_data.get('permit_number', ''),
        # 신청인 업체명 - top-left of L4:V4 merged range
        'L4': form_data.get('company_name', ''),
        # 신청인 직책 - top-left of Z4:AD4 merged range
        'Z4': form_data.get('worker_position', ''),
        # 신청인 성명(실제 서식엔 AL4:BD4가 "성명+(서명)" 입력칸이고, AE4:AK4는 "성명"이라는 고정 라벨임.
        # AE4에 값을 쓰면 라벨을 지워버리는 문제가 있어 AL4 한 곳에서만 "성명 + (서명)"으로 처리함(아래).
        # 작업위치(작업장소) - top-left of R6:AD7 merged range (I6는 "작업위치" 라벨이라 쓰면 안 됨)
        'R6': form_data.get('work_location', ''),
        # 작업대상 - top-left of AL6:BD7 merged range (AE6:AK7은 "작업대상" 고정 라벨)
        'AL6': form_data.get('work_target', ''),
        'I8': form_data.get('work_description', ''),  # 작업 개요 (top-left of I8:BD9 merged range)
        'I36': form_data.get('special_notes', ''),  # 기타 특별사항 (top-left of I36:BD36 merged range)
        'L40': completion_time_value,  # 작업완료 시간: 작업허가기간의 종료시간 (top-left of L40:S40 merged range)
        'AI40': worker_count_value,  # 작업자 명수 (top-left of AI40:AO40 merged range)
    }
    
    # 작업허가기간 = 항상 작업일자 기준 (년/월/일), 시작~종료 시간 별도 입력
    work_date_str = form_data.get('work_date', '')
    if work_date_str:
        try:
            _d = datetime.strptime(work_date_str, '%Y-%m-%d')
            # 템플릿의 날짜 입력칸: I5:M5(년), P5:R5(월), U5:W5(일)
            cell_mapping['I5'] = str(_d.year)
            cell_mapping['P5'] = str(_d.month)
            cell_mapping['U5'] = str(_d.day)
            # 허가일자(AE3 라벨 옆 AL3:BD3 입력칸)는 작업허가기간과 별개 칸이라 따로 채워야 함.
            # 앱에 별도 "허가일자" 입력이 없어 우선 작업일자와 동일하게 자동 기입함.
            cell_mapping['AL3'] = f'{_d.year}년 {_d.month}월 {_d.day}일'
        except ValueError:
            pass
    # 템플릿은 Z5:AB5(시부터 앞)와 AH5:AJ5(시까지 앞)에 시간을 기입한다.
    # "시 부터/시까지" 라벨 자체가 "시(hour)"만 의미하므로 시(HH)만 기입한다.
    if form_data.get('permit_start_time'):
        try:
            _h, _m = form_data['permit_start_time'].split(':')[:2]
            cell_mapping['Z5'] = _h  # 시 부터 (시)
        except ValueError:
            pass
    if form_data.get('permit_end_time'):
        try:
            _h, _m = form_data['permit_end_time'].split(':')[:2]
            cell_mapping['AH5'] = _h  # 시까지 (시)
        except ValueError:
            pass

    # 화면에서 선택한 작업종류에 맞춰 Excel의 해당 작업 대표 체크칸을 표시한다.
    selected_work_types = form_data.get('work_types') or []
    if isinstance(selected_work_types, str):
        selected_work_types = [item.strip() for item in selected_work_types.split(',') if item.strip()]
    if not selected_work_types:
        work_type_text = form_data.get('work_type', '')
        selected_work_types = [
            work_type_name
            for work_type_name in WORK_TYPE_CHECKBOX_CELL_MAP
            if work_type_name in work_type_text
        ]
    for work_type_name in selected_work_types:
        target_cell = WORK_TYPE_CHECKBOX_CELL_MAP.get(work_type_name)
        if target_cell:
            cell_mapping[target_cell] = '☑'
    
    # 체크된 항목만 표시(☑). 체크 안 된 항목은 서식 원본의 '□'를 그대로 둔다.
    for (category, item), cell in CHECKBOX_CELL_MAP.items():
        if safety_data.get(category, {}).get(item, False):
            cell_mapping[cell] = '☑'
    
    # 밀폐공간/정전 카테고리 전체("해당사항") 체크박스
    for category, cell in CATEGORY_CHECKBOX_CELL_MAP.items():
        if safety_data.get(category, {}).get('해당', False):
            cell_mapping[cell] = '☑'
    
    # 정전: 선택한 위치(제어실/현장)에 맞는 셀에 체크
    power_outage_location = form_data.get('power_outage_location', '')
    if power_outage_location in POWER_OUTAGE_LOCATION_CELLS:
        location_cells = POWER_OUTAGE_LOCATION_CELLS[power_outage_location]
        power_outage_checks = safety_data.get('정전', {})
        for item, cell in location_cells.items():
            if power_outage_checks.get(item, False):
                cell_mapping[cell] = '☑'
    
    # 위험성평가 필요 여부 (작업절차서변화 / 작업상이) 유·무 체크
    for field_key, choices in RISK_ASSESSMENT_CELLS.items():
        selected = form_data.get(field_key)
        if selected in choices:
            cell_mapping[choices[selected]] = '☑'
    
    # 텍스트 입력 항목
    for field_name, cell in TEXT_FIELD_CELL_MAP.items():
        value = text_field_data.get(field_name, '')
        if value:
            cell_mapping[cell] = value

    # 밀폐공간 가스농도 측정 4건을 서식의 22~23행 좌/우 표에 매핑한다.
    for measurement, target_cells in zip(
        form_data.get('gas_measurements', []), GAS_MEASUREMENT_CELL_MAP
    ):
        if not isinstance(measurement, dict):
            continue
        for field_name, cell in target_cells.items():
            value = measurement.get(field_name, '')
            if value:
                cell_mapping[cell] = value
    
    # 신청인란(AL4)과 시행업체 책임자란(T37)에 동일하게 "성명 + (서명)"을 넣는다(같은 사람, 같은 이름).
    # 서명은 책임자 서명 1개만 받으므로, 이미지는 아래 signatures 처리 블록에서 이 두 칸에 동일하게 겹쳐서 삽입됨.
    worker_name_for_sig = form_data.get('worker_name', '')
    if worker_name_for_sig:
        cell_mapping['AL4'] = f'{worker_name_for_sig}          (서명)'
        cell_mapping['T37'] = f'{worker_name_for_sig}          (서명)'

    # 담당자(관리자, 관리자 페이지에서 지정) - 입회자 / 발급자(담당자) / 각 작업별 확인자란에 이름을 넣는다.
    # 입회자는 이름만, 발급자/확인자는 서식에 맞춰 "(서명)" 문구를 함께 넣는다.
    manager_name = form_data.get('manager_name', '')
    if manager_name:
        manager_attendee_text = manager_name
        manager_sig_text = f'{manager_name}     (서명)'
        confined_space_sig_text = f'{manager_name} (서명)'
        cell_mapping['AI37'] = manager_sig_text  # 37행 입회자 (top-left of AI37:AV37)
        cell_mapping['W38'] = manager_sig_text  # 발급자(담당자) (top-left of W38:AD38)
        cell_mapping['X40'] = manager_attendee_text  # 작업완료 40행 입회자 (top-left of X40:AD40)
        cell_mapping['AW41'] = manager_sig_text  # 오른쪽 맨아래 발급자(담당자) 서명칸 (AW41:BD41)

        # 밀폐공간/정전/굴착/고소/중장비가 해당되면 각 확인자란에도 담당자 이름을 넣는다.
        if safety_data.get('밀폐공간', {}).get('해당', False):
            cell_mapping['AX19'] = confined_space_sig_text  # 밀폐공간 확인자
        if safety_data.get('정전', {}).get('해당', False):
            cell_mapping['AT26'] = manager_sig_text  # 정전(현장) 확인자
            cell_mapping['AT28'] = manager_sig_text  # 정전(전원복구) 확인자
        if '굴착' in safety_data:
            cell_mapping['AT30'] = manager_sig_text  # 굴착 확인자
        if '고소' in safety_data:
            cell_mapping['AT32'] = manager_sig_text  # 고소 확인자
        if '중장비' in safety_data:
            cell_mapping['AT35'] = manager_sig_text  # 중장비 확인자

    # 승인자(팀장, 관리자 페이지에서 지정) - 서명 이미지 없이 이름 텍스트만 넣는다.
    approver_name = form_data.get('approver_name', '')
    if approver_name:
        cell_mapping['W39'] = f'{approver_name}     (서명)'  # 승인자(팀장) 성명 (top-left of W39:AD39)


    # Write field values
    for cell, value in cell_mapping.items():
        # Skip if value is empty
        if not value:
            continue
            
        try:
            # Check if cell is merged and write to top-left cell
            from openpyxl.cell import MergedCell
            if isinstance(ws[cell], MergedCell):
                # Find the merged range and write to top-left cell
                for merged_range in ws.merged_cells.ranges:
                    if cell in merged_range:
                        top_left = merged_range.start_cell.coordinate
                        ws[top_left].value = value
                        break
            else:
                ws[cell] = value
        except Exception as e:
            print(f"Error writing to cell {cell}: {e}")
            continue
    
    # Handle signature image if available
    # 책임자 서명 하나만 받아서, 신청인(AL4)란과 시행업체 책임자(T37)란 두 곳에 동일하게 삽입한다.
    # "(서명)" 안내 문구는 위(셀 상단)에 오도록 정렬하고, 서명 이미지는 그 아래로 오프셋을 줘서 겹치지 않게 한다.
    if 'signatures' in form_data and form_data['signatures']:
        try:
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            from openpyxl.utils.cell import coordinate_from_string
            
            signature_target_cells = ['T37', 'AL4']  # 시행업체(책임자) 서명란, 신청인 서명란
            sig_key = 'company_rep'
            sig_data = form_data['signatures'].get(sig_key)
            
            if sig_data is not None:
                try:
                    # Handle both list and numpy array data
                    img_array = sig_data
                    if isinstance(img_array, list):
                        import numpy as np
                        img_array = np.array(img_array)
                    elif hasattr(img_array, 'tolist'):
                        img_array = img_array.tolist()
                        import numpy as np
                        img_array = np.array(img_array)
                    
                    img = PILImage.fromarray(img_array.astype('uint8'))
                    
                    # Remove white background (make transparent)
                    img = img.convert('RGBA')
                    datas = img.getdata()
                    new_data = []
                    for item in datas:
                        # If pixel is white or near white, make it transparent
                        if item[0] > 200 and item[1] > 200 and item[2] > 200:
                            new_data.append((255, 255, 255, 0))
                        else:
                            new_data.append(item)
                    img.putdata(new_data)
                    
                    # Resize image to fit within a single ~20px-tall row (텍스트와 세로로 겹치지 않게 작게)
                    img = img.resize((60, 16))
                    
                    # Save to bytes once, reused for both insertions
                    img_bytes = io.BytesIO()
                    img.save(img_bytes, format='PNG')
                    png_data = img_bytes.getvalue()
                    
                    for cell_pos in signature_target_cells:
                        try:
                            # 이미지를 끝에서 네 열 앞에 둬 오른쪽 인접 칸으로 넘어가지 않게 한다.
                            # 열 안에서만 offset을 사용해 Excel에서의 위치 해석도 안정적으로 유지한다.
                            _, row_number = coordinate_from_string(cell_pos)
                            merged_range = next(
                                (r for r in ws.merged_cells.ranges if cell_pos in r),
                                None,
                            )
                            if merged_range is None:
                                raise ValueError(f"Signature cell {cell_pos} is not merged")

                            img_w_px, img_h_px = 60, 16
                             
                            marker = AnchorMarker(
                                col=merged_range.max_col - 5,
                                colOff=pixels_to_EMU(25),
                                row=row_number - 1,
                                rowOff=pixels_to_EMU(2),
                            )
                            size = XDRPositiveSize2D(pixels_to_EMU(img_w_px), pixels_to_EMU(img_h_px))
                            anchor = OneCellAnchor(_from=marker, ext=size)
                            
                            img_obj = OpenpyxlImage(io.BytesIO(png_data))
                            img_obj.anchor = anchor
                            ws.add_image(img_obj)
                        except Exception as e:
                            print(f"Error placing signature at {cell_pos}: {e}")
                            continue
                except Exception as e:
                    print(f"Error processing signature {sig_key}: {e}")
        except Exception as e:
            print(f"Error handling signatures: {e}")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Custom CSS
st.markdown("""
<style>
    .main {
        padding: 1rem;
    }
    .stButton>button {
        width: 100%;
        margin: 0.5rem 0;
    }
    .section-header {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        font-weight: bold;
    }
    .checkbox-group {
        background-color: #fafafa;
        padding: 1rem;
        border-radius: 8px;
        margin: 0.5rem 0;
    }
    .category-title {
        background-color: #f8fafc;
        border-left: 4px solid #2563eb;
        border-radius: 4px;
        margin: 1rem 0 0.35rem;
        padding: 0.65rem 0.85rem;
        font-weight: 700;
    }
    .basic-info-marker {
        display: none;
    }
    /* 기본 정보의 두 영역을 서로 구분되는 카드로 표시한다. */
    [data-testid="stHorizontalBlock"]:has(.basic-info-marker) > [data-testid="column"] {
        background: #f8fafc;
        border: 1px solid #dbe4f0;
        border-radius: 10px;
        padding: 0.75rem 1rem;
    }
    @media (max-width: 640px) {
        .block-container {
            padding: 0.75rem;
        }
        [data-testid="stHorizontalBlock"] {
            flex-direction: column;
            gap: 0;
        }
        [data-testid="stHorizontalBlock"] > [data-testid="column"] {
            min-width: 100% !important;
            width: 100% !important;
            flex: 1 1 100% !important;
        }
    }
</style>
""", unsafe_allow_html=True)

# Page navigation
page = st.sidebar.radio("페이지 선택", ["👷 현장 작업자", "👨‍💼 관리자"])

if page == "👷 현장 작업자":
    st.title("🛡️ 안전작업허가서 - 현장 작업자")
    st.markdown("---")

    # Initialize session state for signatures
    if 'signatures' not in st.session_state:
        st.session_state.signatures = {}

    # Section 1: 기본 정보
    st.markdown('<div class="section-header">📋 기본 정보</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<span class="basic-info-marker"></span>', unsafe_allow_html=True)
        st.markdown("#### 작업 정보")
        work_date = st.date_input("작업일자", datetime.now(), key="work_date")
        permit_start_time = st.text_input(
            "작업 시작 시간", key="permit_start_time_text", placeholder="예: 8 또는 0830",
            help="숫자만 입력하면 자동으로 HH:MM 형식으로 바뀝니다. 예: 8 → 08:00, 1730 → 17:30",
            max_chars=5, on_change=normalize_time_widget, args=("permit_start_time_text",),
        )
        if st.session_state.get("permit_start_time_text_error"):
            st.error(f"작업 시작 시간: {st.session_state['permit_start_time_text_error']}")
        permit_end_time = st.text_input(
            "작업 종료 시간", key="permit_end_time_text", placeholder="예: 17 또는 1730",
            help="숫자만 입력하면 자동으로 HH:MM 형식으로 바뀝니다. 예: 17 → 17:00, 1730 → 17:30",
            max_chars=5, on_change=normalize_time_widget, args=("permit_end_time_text",),
        )
        if st.session_state.get("permit_end_time_text_error"):
            st.error(f"작업 종료 시간: {st.session_state['permit_end_time_text_error']}")
        work_location = st.text_input("작업장소", key="work_location")
        work_target = st.text_input("작업대상", key="work_target")
        work_types = st.multiselect(
            "작업종류 (복수 선택 가능)",
            ["굴착작업", "고소작업", "중장비작업", "전기작업", "용접작업", "밀폐작업", "기타"],
            key="work_types",
            help="선택한 모든 작업 종류의 점검 항목이 아래에 표시됩니다.",
        )
        work_type = ", ".join(work_types) if work_types else "선택 안 함"

    with col2:
        st.markdown("#### 신청 정보")
        construction_name = st.text_input("공사명", key="construction_name")
        # 허가번호 = 공사명 + 순번(같은 공사명으로 이미 제출된 건수+1, 2자리)
        _existing_forms = load_submitted_forms()
        _seq = sum(1 for f in _existing_forms if f.get('construction_name') == construction_name) + 1 if construction_name else 0
        permit_number = f"{construction_name}{_seq:02d}" if construction_name else ""
        if construction_name:
            st.caption(f"허가번호 미리보기: {permit_number}")
        company_name = st.text_input("업체명", key="company_name")
        worker_position = st.text_input("직책", key="worker_position")
        worker_name = st.text_input("성명", key="worker_name")
        worker_count = st.number_input(
            "작업자(명)",
            min_value=0,
            step=1,
            value=0,
            key="worker_count",
            help="작업 인원을 숫자로 입력하세요.",
        )
        # 담당자(관리자)는 현장 작업자가 아니라 관리자 페이지에서 제출 건별로 지정한다.

    # Section 2: 작업 내용
    st.markdown('<div class="section-header">📝 작업 내용</div>', unsafe_allow_html=True)
    work_description = st.text_area("작업 내용 설명", height=100, placeholder="작업의 구체적인 내용을 입력하세요...", key="work_description")

    # Section 3: 서명
    # 모바일에서는 작업종류별 안전조치가 길어지므로, 서명은 조건부 항목보다 먼저 받는다.
    st.markdown('<div class="section-header">✍️ 서명</div>', unsafe_allow_html=True)

    def signature_canvas(key, label):
        st.markdown(f"**{label}**")
        canvas_result = st_canvas(
            fill_color="rgba(255, 165, 0, 0.3)",
            stroke_width=3,
            stroke_color="#000000",
            background_color="#ffffff",
            background_image=None,
            update_streamlit=True,
            height=150,
            width=320,
            drawing_mode="freedraw",
            point_display_radius=0,
            key=key,
        )
        if canvas_result.image_data is not None:
            st.session_state.signatures[key] = canvas_result.image_data
        return canvas_result

    # 발급자/승인자/입회자/작업자 서명은 관리자가 나중에 처리하므로 현장 작업자 화면에서는 받지 않음
    # 책임자 서명 하나만 받고, 신청인란·시행업체(책임자)란 둘 다 이 서명 + 위에서 입력한 성명으로 채움
    signature_canvas("company_rep", "책임자 서명")

    # Section 4: 안전조치 확인 (Checkboxes)
    st.markdown('<div class="section-header">✅ 안전조치 확인</div>', unsafe_allow_html=True)

    # Safety check categories (matching Excel template's actual checkbox cells)
    # 주의: "차단기기"/"설비"는 실제 서식엔 채울 칸이 없는 고정 라벨이라 제외함
    # 밀폐공간/정전은 카테고리 해당여부 + 위치선택이 있어 아래에서 별도 UI로 렌더링함(이 dict엔 없음)
    safety_checks = {
        "첨부서류": [
            "작업계획서",
            "기술자료(도면)",
            "소화기목록",
            "안전장구 목록",
            "특수작업절차서",
            "굴착도면"
        ],
        "안전조치 요구사항": [
            "작업구역 설정(출입경고 표지)",
            "작업구역 가연성물질 제거",
            "밸브차단 및 차단표지부착(도면 비교)",
            "맹판설치 및 표지부착(도면 비교)",
            "위험물질(가연성분진 포함)방출 및 처리",
            "보충작업허가",  # 서식에 체크박스는 없지만 목록엔 남겨둠(엑셀 미반영)
            "용기개방 및 압력방출",
            "용기내부 세정 및 처리",
            "불활성가스 치환 및 환기",
            "환기장비",
            "가스농도 측정",
            "조명장비",
            "소화기",
            "안전장구",
            "안전교육",
            "운전요원의 입회"
        ],
        "굴착": [
            "가스,기계,소방배관",
            "전기,계장,통신"
        ],
        "고소": [
            "작업발판, 안전난간",
            "안전사다리 사용",
            "안전대 착용·부착",
            "추락방지망"
        ],
        "중장비": [
            "기상, 노면상태",
            "자격증소지",
            "현장책임자 감독",
            "전선, 설비 간섭",
            "신호수배치",
            "매트 등 부속장구"
        ]
    }

    # 체크박스가 아니라 텍스트로 채우는 항목 (실제 서식엔 빈칸으로 존재)
    safety_text_fields = {
        "굴착": [
            "굴착 가스,기계,소방배관 점검자",
            "굴착 전기,계장,통신 점검자",
            "굴착 허가기간"
        ],
        "고소": [
            "고소 허가기간"
        ],
        "중장비": [
            "투입장비명",
            "운전원 성명",
            "중장비 허가기간"
        ]
    }

    # 공통 항목은 항상 보이고, 작업종류별 항목은 선택한 작업에만 표시한다.
    work_type_categories = {
        "굴착작업": ["굴착"],
        "고소작업": ["고소"],
        "중장비작업": ["중장비"],
    }
    visible_safety_categories = ["첨부서류", "안전조치 요구사항"]
    required_safety_checks = {("안전조치 요구사항", "안전장구"), ("안전조치 요구사항", "안전교육")}
    for selected_work_type in work_types:
        for category in work_type_categories.get(selected_work_type, []):
            if category not in visible_safety_categories:
                visible_safety_categories.append(category)

    def render_safety_category(category, checks, fields, column_count=2):
        """카테고리별 체크/입력 항목을 넓은 화면에서는 여러 열로 표시한다."""
        st.markdown(f'<div class="category-title">{category}</div>', unsafe_allow_html=True)
        if category == "안전조치 요구사항":
            st.info("필수 확인: 안전장구와 안전교육을 모두 체크해야 제출할 수 있습니다.")

        checkbox_columns = st.columns(min(column_count, len(checks)))
        for index, check in enumerate(checks):
            with checkbox_columns[index % len(checkbox_columns)]:
                label = f"{check} (필수)" if (category, check) in required_safety_checks else check
                st.checkbox(label, key=f"check_{category}_{check}")

        if fields:
            text_columns = st.columns(min(2, len(fields)))
            for index, field in enumerate(fields):
                with text_columns[index % len(text_columns)]:
                    if "허가기간" in field:
                        st.markdown(f"**{field}**")
                        st.info(f"{work_date} (작업일자와 동일)")
                    else:
                        st.text_input(field, key=f"text_{category}_{field}")

    for category in visible_safety_categories:
        checks = safety_checks[category]
        # 긴 안전조치 문구는 두 열로, 나머지는 공간을 활용해 세 열로 나눈다.
        column_count = 2 if category == "안전조치 요구사항" else 3
        render_safety_category(
            category,
            checks,
            safety_text_fields.get(category, []),
            column_count,
        )

    # 위험성평가 필요 여부 (작업절차서변화 / 작업상이 각각 유·무 선택)
    st.markdown('<div class="checkbox-group"><strong>위험성평가 필요 여부</strong></div>', unsafe_allow_html=True)
    risk_assessment_change = st.radio("작업절차서변화", ["선택 안 함", "유", "무"], key="risk_assessment_change", horizontal=True)
    risk_assessment_diff = st.radio("작업상이", ["선택 안 함", "유", "무"], key="risk_assessment_diff", horizontal=True)

    # 밀폐공간은 밀폐작업 선택 시에만 표시한다.
    is_confined_space_work = "밀폐작업" in work_types
    if is_confined_space_work:
        st.markdown('<div class="checkbox-group"><strong>밀폐공간</strong></div>', unsafe_allow_html=True)
        confined_space_applies = st.checkbox("밀폐공간 작업 해당", key="check_밀폐공간_해당")
        st.checkbox("통신수단", key="check_밀폐공간_통신수단")
        st.checkbox("구명장구(줄, 송기마스크)", key="check_밀폐공간_구명장구(줄, 송기마스크)")
        st.markdown("**밀폐공간 허가기간**")
        st.info(f"{work_date} (작업일자와 동일)")
        st.caption("참고자료: 가스농도 측정결과 1. HC: 0%, 2. O2: 18%이상, 3. CO: 30ppm미만, 4. CO2: 1.5%미만, 5. H2S: 10ppm미만")

        if confined_space_applies:
            st.markdown("#### 🧪 가스농도 측정")
            st.caption("Excel 서식의 22~23행 좌·우 가스농도 측정란에 각각 입력됩니다.")
            for measurement_index in range(1, 5):
                st.markdown(f"**측정 {measurement_index}**")
                gas_col1, gas_col2, gas_col3, gas_col4 = st.columns([2, 1.5, 1.5, 2])
                with gas_col1:
                    st.text_input(
                        "물질명",
                        key=f"gas_{measurement_index}_material_name",
                        placeholder="예: O2, CO, H2S",
                    )
                with gas_col2:
                    st.text_input(
                        "결과",
                        key=f"gas_{measurement_index}_result",
                        placeholder="예: 20.9%, 0ppm",
                    )
                with gas_col3:
                    st.time_input(
                        "측정시간",
                        value=None,
                        key=f"gas_{measurement_index}_measurement_time",
                    )
                with gas_col4:
                    st.text_input(
                        "측정자/확인자",
                        key=f"gas_{measurement_index}_measurer_confirmer",
                        placeholder="성명 입력",
                    )

    # 정전 항목은 전기작업 선택 시에만 표시한다.
    is_power_outage_work = "전기작업" in work_types
    power_outage_location = "선택 안 함"
    if is_power_outage_work:
        st.markdown('<div class="checkbox-group"><strong>정전</strong></div>', unsafe_allow_html=True)
        power_outage_applies = st.checkbox("정전 작업 해당", key="check_정전_해당")
        power_outage_location = st.radio(
            "차단 위치", ["선택 안 함", "제어실", "현장"], key="power_outage_location", horizontal=True
        )
        st.checkbox("스위치/차단기 내림", key="check_정전_스위치/차단기 내림")
        st.checkbox("잠금장치 시건, 표지부착", key="check_정전_잠금장치 시건, 표지부착")
        st.markdown("**정전 허가기간**")
        st.info(f"{work_date} (작업일자와 동일)")
        st.info("전원복구 참고: 모든 작업이 완료 된 후 운전부서의 입회자의 요청에 의해서만 전원을 복구하여야 한다.")
        st.text_input("전원복구 요청자", key="text_정전_전원복구 요청자")
        st.text_input("전원복구 시간", key="text_정전_전원복구 시간")

    # Additional safety notes
    special_notes = st.text_input("기타 특별사항", key="special_notes")

    # Section 5: 작업 허가 연장
    st.markdown('<div class="section-header">⏰ 작업 허가 연장</div>', unsafe_allow_html=True)

    extend_requested = st.radio(
        "연장 신청 여부", ["신청 안 함", "연장 신청"], key="extend_requested", horizontal=True
    )
    extend_date = extend_start = extend_end = None
    if extend_requested == "연장 신청":
        col1, col2, col3 = st.columns(3)
        with col1:
            extend_date = st.date_input("연장일자", key="extend_date")
        with col2:
            extend_start = st.time_input("연장 시작 시간", key="extend_start")
        with col3:
            extend_end = st.time_input("연장 종료 시간", key="extend_end")

    # Action buttons
    st.markdown("---")
    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("📤 제출하기", use_container_width=True, type="primary"):
            normalized_start, start_error = normalize_clock_time(permit_start_time)
            normalized_end, end_error = normalize_clock_time(permit_end_time)
            missing_required_checks = [
                check
                for category, check in required_safety_checks
                if not st.session_state.get(f"check_{category}_{check}", False)
            ]

            validation_errors = []
            if start_error:
                validation_errors.append(f"작업 시작 시간: {start_error}")
            if end_error:
                validation_errors.append(f"작업 종료 시간: {end_error}")
            if missing_required_checks:
                validation_errors.append(
                    "안전조치 요구사항의 필수 항목을 체크해 주세요: "
                    + ", ".join(sorted(missing_required_checks))
                )

            if validation_errors:
                for error in validation_errors:
                    st.error(error)
                st.stop()

            # Collect form data
            form_data = {
                'id': datetime.now().strftime("%Y%m%d%H%M%S"),
                'submitted_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'work_date': str(work_date),
                'permit_start_time': normalized_start or None,
                'permit_end_time': normalized_end or None,
                'work_location': work_location,
                'work_target': work_target,
                'work_type': work_type,
                'work_types': work_types,
                'construction_name': construction_name,
                'permit_number': permit_number,
                'company_name': company_name,
                'worker_position': worker_position,
                'worker_name': worker_name,
                'worker_count': int(worker_count) if worker_count else None,
                'manager_name': None,  # 담당자(관리자)는 관리자 페이지에서 나중에 지정
                'work_description': work_description,
                'special_notes': special_notes,
                'power_outage_location': None if power_outage_location == "선택 안 함" else power_outage_location,
                'risk_assessment_change': None if risk_assessment_change == "선택 안 함" else risk_assessment_change,
                'risk_assessment_diff': None if risk_assessment_diff == "선택 안 함" else risk_assessment_diff,
                'extend_requested': extend_requested == "연장 신청",
                'extend_date': str(extend_date) if extend_date else None,
                'extend_start': str(extend_start) if extend_start else None,
                'extend_end': str(extend_end) if extend_end else None,
                'safety_checks': {},
                'safety_text_fields': {},
                'gas_measurements': [],
                'signatures': st.session_state.signatures
            }
            
            # Collect safety checks
            for category in visible_safety_categories:
                checks = safety_checks[category]
                form_data['safety_checks'][category] = {}
                for check in checks:
                    form_data['safety_checks'][category][check] = st.session_state.get(f"check_{category}_{check}", False)
            
            # Collect safety text fields (투입장비명, 운전원 성명, 각종 허가기간/점검자 등)
            for category in visible_safety_categories:
                fields = safety_text_fields.get(category, [])
                for field in fields:
                    if "허가기간" in field:
                        form_data['safety_text_fields'][field] = str(work_date)
                    else:
                        value = st.session_state.get(f"text_{category}_{field}", "")
                        if value:
                            form_data['safety_text_fields'][field] = str(value)
            
            # 밀폐공간 (밀폐작업일 때만 저장)
            if is_confined_space_work:
                form_data['safety_checks']['밀폐공간'] = {
                    '해당': st.session_state.get("check_밀폐공간_해당", False),
                    '통신수단': st.session_state.get("check_밀폐공간_통신수단", False),
                    '구명장구(줄, 송기마스크)': st.session_state.get("check_밀폐공간_구명장구(줄, 송기마스크)", False),
                }
                form_data['safety_text_fields']['밀폐공간 허가기간'] = str(work_date)

                if confined_space_applies:
                    for measurement_index in range(1, 5):
                        measurement_time = st.session_state.get(
                            f"gas_{measurement_index}_measurement_time"
                        )
                        form_data['gas_measurements'].append({
                            'material_name': st.session_state.get(
                                f"gas_{measurement_index}_material_name", ""
                            ).strip(),
                            'result': st.session_state.get(
                                f"gas_{measurement_index}_result", ""
                            ).strip(),
                            'measurement_time': (
                                measurement_time.strftime('%H:%M') if measurement_time else ""
                            ),
                            'measurer_confirmer': st.session_state.get(
                                f"gas_{measurement_index}_measurer_confirmer", ""
                            ).strip(),
                        })
            
            # 정전 (전기작업일 때만 저장)
            if is_power_outage_work:
                form_data['safety_checks']['정전'] = {
                    '해당': st.session_state.get("check_정전_해당", False),
                    '스위치/차단기 내림': st.session_state.get("check_정전_스위치/차단기 내림", False),
                    '잠금장치 시건, 표지부착': st.session_state.get("check_정전_잠금장치 시건, 표지부착", False),
                }
                form_data['safety_text_fields']['정전 허가기간'] = str(work_date)
                for field in ["전원복구 요청자", "전원복구 시간"]:
                    value = st.session_state.get(f"text_정전_{field}", "")
                    if value:
                        form_data['safety_text_fields'][field] = str(value)
            
            # Save form
            if save_form(form_data):
                st.success("✅ 안전작업허가서가 제출되었습니다!")
                st.balloons()

    with col2:
        if st.button("🔄 초기화", use_container_width=True):
            st.rerun()

    with col3:
        if st.button("📄 미리보기", use_container_width=True):
            st.info("제출 후 관리자 페이지에서 Excel로 출력하세요.")

    # Summary section
    st.markdown("---")
    st.markdown('<div class="section-header">📊 입력 내용 요약</div>', unsafe_allow_html=True)

    summary_data = {
        "항목": ["작업일자", "작업 시작~종료", "작업장소", "작업대상", "공사명(허가번호)", "작업종류", "업체명", "성명", "직책", "작업자(명)"],
        "내용": [
            work_date,
            f"{permit_start_time or '-'} ~ {permit_end_time or '-'}",
            work_location,
            work_target,
            permit_number,
            work_type,
            company_name,
            worker_name,
            worker_position,
            worker_count if worker_count else ""
        ]
    }

    st.table(pd.DataFrame(summary_data))

    # Display checked safety items
    st.markdown("### ✅ 확인된 안전조치 항목")
    checked_items = []
    for category in visible_safety_categories:
        checks = safety_checks[category]
        for check in checks:
            if st.session_state.get(f"check_{category}_{check}", False):
                checked_items.append(f"{category}: {check}")
    if is_confined_space_work:
        for check in ["해당", "통신수단", "구명장구(줄, 송기마스크)"]:
            if st.session_state.get(f"check_밀폐공간_{check}", False):
                checked_items.append(f"밀폐공간: {check}")
    if is_power_outage_work:
        for check in ["해당", "스위치/차단기 내림", "잠금장치 시건, 표지부착"]:
            if st.session_state.get(f"check_정전_{check}", False):
                checked_items.append(f"정전: {check}")

    if checked_items:
        for item in checked_items:
            st.markdown(f"- {item}")
    else:
        st.info("아직 확인된 안전조치 항목이 없습니다.")

    # Display filled text fields (허가기간, 점검자, 전원복구, 투입장비 등)
    filled_text_fields = []
    for category in visible_safety_categories:
        fields = safety_text_fields.get(category, [])
        for field in fields:
            if "허가기간" in field:
                filled_text_fields.append(f"{field}: {work_date}")
            else:
                value = st.session_state.get(f"text_{category}_{field}", "")
                if value:
                    filled_text_fields.append(f"{field}: {value}")
    if is_confined_space_work:
        filled_text_fields.append(f"밀폐공간 허가기간: {work_date}")
    if is_power_outage_work:
        filled_text_fields.append(f"정전 허가기간: {work_date}")
        for field in ["전원복구 요청자", "전원복구 시간"]:
            value = st.session_state.get(f"text_정전_{field}", "")
            if value:
                filled_text_fields.append(f"{field}: {value}")
    if is_power_outage_work and power_outage_location != "선택 안 함":
        filled_text_fields.append(f"정전 차단 위치: {power_outage_location}")

    if filled_text_fields:
        st.markdown("### 📝 입력된 텍스트 항목")
        for item in filled_text_fields:
            st.markdown(f"- {item}")

    if is_confined_space_work and confined_space_applies:
        gas_measurement_summary = []
        for measurement_index in range(1, 5):
            measurement_time = st.session_state.get(f"gas_{measurement_index}_measurement_time")
            measurement = {
                "구분": f"측정 {measurement_index}",
                "물질명": st.session_state.get(f"gas_{measurement_index}_material_name", ""),
                "결과": st.session_state.get(f"gas_{measurement_index}_result", ""),
                "측정시간": measurement_time.strftime('%H:%M') if measurement_time else "",
                "측정자/확인자": st.session_state.get(
                    f"gas_{measurement_index}_measurer_confirmer", ""
                ),
            }
            if any(measurement[field] for field in ["물질명", "결과", "측정시간", "측정자/확인자"]):
                gas_measurement_summary.append(measurement)

        if gas_measurement_summary:
            st.markdown("### 🧪 가스농도 측정 입력 내용")
            st.table(pd.DataFrame(gas_measurement_summary))

else:
    # Admin page
    st.title("👨‍💼 관리자 페이지")
    st.markdown("---")
    
    # Load submitted forms
    forms = load_submitted_forms()
    
    if not forms:
        st.info("📭 제출된 안전작업허가서가 없습니다.")
    else:
        st.markdown(f"### 📋 제출된 양식 ({len(forms)}건)")
        
        # Display forms in a table
        for i, form in enumerate(reversed(forms)):
            with st.expander(f"📄 {form['submitted_at']} - {form['work_location']} ({form['work_type']})"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown(f"**작업일자:** {form['work_date']}")
                    st.markdown(f"**작업 시작~종료:** {form.get('permit_start_time','')} ~ {form.get('permit_end_time','')}")
                    st.markdown(f"**작업장소:** {form['work_location']}")
                    if form.get('work_target'):
                        st.markdown(f"**작업대상:** {form['work_target']}")
                    st.markdown(f"**허가번호(공사명):** {form.get('permit_number','')}")
                    st.markdown(f"**작업종류:** {form['work_type']}")
                    st.markdown(f"**업체명:** {form.get('company_name','')}")
                    st.markdown(f"**성명:** {form['worker_name']}")
                    st.markdown(f"**직책:** {form.get('worker_position','')}")
                    if form.get('worker_count'):
                        st.markdown(f"**작업자(명):** {form.get('worker_count')}")
                
                with col2:
                    if form.get('power_outage_location'):
                        st.markdown(f"**정전 차단 위치:** {form['power_outage_location']}")

                # 담당자(관리자) 지정 - 현장 제출 화면이 아니라 여기서 관리자가 직접 이름을 입력/저장
                st.markdown("**담당자(관리자) 지정:**")
                current_manager = form.get('manager_name') or ""
                col_m1, col_m2 = st.columns([3, 1])
                with col_m1:
                    manager_input = st.text_input(
                        "담당자",
                        value=current_manager,
                        key=f"manager_input_{form['id']}",
                        label_visibility="collapsed",
                        placeholder="담당자 이름 입력",
                    )
                with col_m2:
                    if st.button("저장", key=f"manager_save_{form['id']}", use_container_width=True):
                        value_to_save = manager_input.strip() or None
                        with st.spinner("담당자 저장 중..."):
                            manager_saved = update_manager_name(
                                form['id'], value_to_save, doc_id=form.get('_doc_id')
                            )
                        if manager_saved:
                            form['manager_name'] = value_to_save
                            st.success("담당자가 저장되었습니다.")

                # 승인자(팀장) 지정 - Excel 결과물 생성 시점에 관리자가 직접 이름을 입력/저장
                st.markdown("**승인자(팀장) 지정:**")
                current_approver = form.get('approver_name') or ""
                col_a1, col_a2 = st.columns([3, 1])
                with col_a1:
                    approver_input = st.text_input(
                        "승인자(팀장)",
                        value=current_approver,
                        key=f"approver_input_{form['id']}",
                        label_visibility="collapsed",
                        placeholder="승인자(팀장) 이름 입력",
                    )
                with col_a2:
                    if st.button("저장", key=f"approver_save_{form['id']}", use_container_width=True):
                        value_to_save = approver_input.strip() or None
                        with st.spinner("승인자 저장 중..."):
                            approver_saved = update_form_fields(
                                form['id'], {'approver_name': value_to_save}, doc_id=form.get('_doc_id')
                            )
                        if approver_saved:
                            form['approver_name'] = value_to_save
                            st.success("승인자(팀장)가 저장되었습니다.")

                st.markdown("**작업 내용:**")
                st.text(form['work_description'])
                
                st.markdown("**안전조치 확인:**")
                for category, checks in form['safety_checks'].items():
                    checked = [k for k, v in checks.items() if v]
                    if checked:
                        st.markdown(f"- {category}: {', '.join(checked)}")
                
                text_fields = form.get('safety_text_fields', {})
                if text_fields:
                    st.markdown("**텍스트 입력 항목:**")
                    for field, value in text_fields.items():
                        st.markdown(f"- {field}: {value}")

                gas_measurements = [
                    (index, measurement)
                    for index, measurement in enumerate(form.get('gas_measurements', []), start=1)
                    if isinstance(measurement, dict) and any(measurement.values())
                ]
                if gas_measurements:
                    st.markdown("**가스농도 측정:**")
                    st.table(pd.DataFrame([
                        {
                            "구분": f"측정 {index}",
                            "물질명": measurement.get('material_name', ''),
                            "결과": measurement.get('result', ''),
                            "측정시간": measurement.get('measurement_time', ''),
                            "측정자/확인자": measurement.get('measurer_confirmer', ''),
                        }
                        for index, measurement in gas_measurements
                    ]))
                
                if form['special_notes']:
                    st.markdown(f"**특별사항:** {form['special_notes']}")
                
                # Action buttons for each form
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button(f"📥 Excel 다운로드", key=f"download_{form['id']}", use_container_width=True):
                        excel_data = fill_excel_template(form)
                        if excel_data:
                            st.download_button(
                                label="💾 파일 저장",
                                data=excel_data,
                                file_name=f"안전작업허가서_{form['id']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.error("❌ Excel 템플릿을 찾을 수 없습니다.")
                
                with col2:
                    if st.button(f"🖨️ 인쇄", key=f"print_{form['id']}", use_container_width=True):
                        st.info("📄 Excel 다운로드 후 인쇄하세요.")
                
                with col3:
                    if st.button(f"🗑️ 삭제", key=f"delete_{form['id']}", use_container_width=True):
                        delete_form(form['id'])
                        st.rerun()
        
        # Bulk actions
        st.markdown("---")
        st.markdown("### 🔧 전체 관리")
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("🗑️ 전체 삭제", use_container_width=True):
                st.session_state.confirm_delete_all_forms = True

            if st.session_state.get('confirm_delete_all_forms'):
                st.warning("정말 모든 양식을 삭제하시겠습니까? 이 작업은 되돌릴 수 없습니다.")
                confirm_col, cancel_col = st.columns(2)
                with confirm_col:
                    if st.button("삭제 확정", key="confirm_delete_all_forms_button", use_container_width=True, type="primary"):
                        with st.spinner("전체 양식 삭제 중..."):
                            deleted_count = delete_all_forms()
                        if deleted_count is not None:
                            st.session_state.confirm_delete_all_forms = False
                            st.success(f"{deleted_count}건이 삭제되었습니다.")
                            st.rerun()
                with cancel_col:
                    if st.button("취소", key="cancel_delete_all_forms_button", use_container_width=True):
                        st.session_state.confirm_delete_all_forms = False
                        st.rerun()
        
        with col2:
            if st.button("📊 통계 보기", use_container_width=True):
                st.markdown("### 📊 제출 통계")
                work_types = [f['work_type'] for f in forms]
                type_counts = pd.Series(work_types).value_counts()
                st.bar_chart(type_counts)

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666; font-size: 0.8rem;'>
    안전작업허가서 시스템 | 모바일 친화적 설계
</div>
""", unsafe_allow_html=True)
